#!/usr/bin/env python3
"""
Generate a master Excel spreadsheet (.xlsx) consolidating all LLM math
benchmark results into a single workbook with multiple tabs.

Usage:  python generate_xml_results.py
Output: master_results.xlsx (in the project root)
"""

import json
import os
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))

# ──────────────────────────────────────────────────────────────────────
# Trial registry: (dataset_friendly, model, trial_number, filepath|None)
# 27 slots = 3 datasets x 3 models x 3 trials
# filepath=None means PLACEHOLDER (trial not yet run)
# ──────────────────────────────────────────────────────────────────────
TRIAL_REGISTRY = [
    # --- Advanced Probability & Statistics (1000 questions) ---
    ("Advanced Probability & Statistics", "gemma3:4b", 1,
     "results/advanced_probability_statistics_problems_gemma3_4b_03-27-26T16-09-40.json"),
    ("Advanced Probability & Statistics", "gemma3:4b", 2, None),
    ("Advanced Probability & Statistics", "gemma3:4b", 3, None),
    ("Advanced Probability & Statistics", "phi3:3.8b", 1,
     "results/advanced_probability_statistics_problems_phi3_3_8b_04-02-26T17-31-48.json"),
    ("Advanced Probability & Statistics", "phi3:3.8b", 2, None),
    ("Advanced Probability & Statistics", "phi3:3.8b", 3, None),
    ("Advanced Probability & Statistics", "qwen3:4b", 1, None),
    ("Advanced Probability & Statistics", "qwen3:4b", 2, None),
    ("Advanced Probability & Statistics", "qwen3:4b", 3, None),
    # --- Calculus I (900 questions) ---
    ("Calculus I", "gemma3:4b", 1,
     "results/calculus1_problems_gemma3_4b_03-27-26T13-46-44.json"),
    ("Calculus I", "gemma3:4b", 2,
     "results/calculus1_problems_gemma3_4b_04-02-26T23-40-12.json"),
    ("Calculus I", "gemma3:4b", 3, None),
    ("Calculus I", "phi3:3.8b", 1,
     "results/calculus1_problems_phi3_3_8b_04-02-26T16-22-56.json"),
    ("Calculus I", "phi3:3.8b", 2, None),
    ("Calculus I", "phi3:3.8b", 3, None),
    ("Calculus I", "qwen3:4b", 1, None),
    ("Calculus I", "qwen3:4b", 2, None),
    ("Calculus I", "qwen3:4b", 3, None),
    # --- Grade 8 Math (644 questions) ---
    ("Grade 8 Math", "gemma3:4b", 1,
     "results/grade_8_math_problems_gemma3_4b_03-27-26T18-27-08.json"),
    ("Grade 8 Math", "gemma3:4b", 2, None),
    ("Grade 8 Math", "gemma3:4b", 3, None),
    ("Grade 8 Math", "phi3:3.8b", 1,
     "results/grade_8_math_problems_phi3_3_8b_04-02-26T18-44-43.json"),
    ("Grade 8 Math", "phi3:3.8b", 2, None),
    ("Grade 8 Math", "phi3:3.8b", 3, None),
    ("Grade 8 Math", "qwen3:4b", 1, None),
    ("Grade 8 Math", "qwen3:4b", 2, None),
    ("Grade 8 Math", "qwen3:4b", 3, None),
]

# Question Results columns: (key, description, is_numeric)
QR_COLUMNS = [
    ("dataset",                "Dataset name",                                     False),
    ("model",                  "Model identifier",                                 False),
    ("trial_number",           "Trial number (1-3)",                                True),
    ("question_id",            "Question ID from dataset",                          False),
    ("category",               "Question category/topic",                           False),
    ("question_text",          "Full question text",                                False),
    ("expected_answer",        "Correct answer from dataset",                       False),
    ("alternate_answer",       "Alternate acceptable answer",                       False),
    ("extracted_answer",       "Answer extracted from LLM response",                False),
    ("is_correct",             "Whether LLM answer was judged correct",             False),
    ("success",                "Whether LLM produced a response",                   False),
    ("processing_time_s",      "Total processing time in seconds",                  True),
    ("extraction_method",      "Method used to extract answer",                     False),
    ("extraction_confidence",  "Confidence of answer extraction (0-1)",             True),
    ("match_type",             "Type of answer match (exact/equivalent/no_match)",  False),
    ("gpu_util_avg_pct",       "Average GPU utilization percent",                   True),
    ("gpu_vram_avg_mb",        "Average GPU VRAM usage in MB",                      True),
    ("gpu_temp_avg_c",         "Average GPU temperature in Celsius",                True),
    ("gpu_power_avg_w",        "Average GPU power draw in Watts",                   True),
    ("energy_wh",              "Estimated energy consumption in Wh",                True),
    ("total_duration_s",       "Total Ollama duration in seconds",                  True),
    ("prompt_eval_count",      "Number of prompt tokens evaluated",                 True),
    ("eval_count",             "Number of output tokens generated",                 True),
    ("generation_speed_tps",   "Token generation speed (tokens/sec)",               True),
]

# JSON source path for each column (used in Data Dictionary)
QR_SOURCE_MAP = {
    "dataset":                "metadata.dataset_file (mapped to friendly name)",
    "model":                  "metadata.model_name",
    "trial_number":           "Assigned chronologically per model+dataset",
    "question_id":            "results[].question_id",
    "category":               "results[].category",
    "question_text":          "results[].question_text",
    "expected_answer":        "results[].expected_answer",
    "alternate_answer":       "results[].alternate_answer",
    "extracted_answer":       "results[].verification.extracted_answer",
    "is_correct":             "results[].is_correct",
    "success":                "results[].success",
    "processing_time_s":      "results[].processing_time",
    "extraction_method":      "results[].verification.extraction_method",
    "extraction_confidence":  "results[].verification.extraction_confidence",
    "match_type":             "results[].verification.match_type",
    "gpu_util_avg_pct":       "results[].system_metrics.gpu_util_avg_percent",
    "gpu_vram_avg_mb":        "results[].system_metrics.gpu_vram_avg_mb",
    "gpu_temp_avg_c":         "results[].system_metrics.gpu_temp_avg_c",
    "gpu_power_avg_w":        "results[].system_metrics.gpu_power_avg_w",
    "energy_wh":              "results[].system_metrics.energy_estimate_wh",
    "total_duration_s":       "results[].ollama_metrics.total_duration_s",
    "prompt_eval_count":      "results[].ollama_metrics.prompt_eval_count",
    "eval_count":             "results[].ollama_metrics.eval_count",
    "generation_speed_tps":   "results[].ollama_metrics.generation_speed_tps",
}

# Styling constants
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
PLACEHOLDER_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


# ──────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────

def load_json(filepath):
    """Load a JSON results file from a path relative to PROJECT_ROOT."""
    abs_path = os.path.join(PROJECT_ROOT, filepath)
    with open(abs_path, "r", encoding="utf-8") as f:
        return json.load(f)


def extract_question_row(dataset, model, trial_num, q):
    """Extract a flat dict of column values from one question result object."""
    v = q.get("verification") or {}
    sm = q.get("system_metrics") or {}
    om = q.get("ollama_metrics") or {}

    is_correct_val = q.get("is_correct")
    success_val = q.get("success")

    return {
        "dataset":                dataset,
        "model":                  model,
        "trial_number":           trial_num,
        "question_id":            q.get("question_id"),
        "category":               q.get("category"),
        "question_text":          q.get("question_text"),
        "expected_answer":        q.get("expected_answer"),
        "alternate_answer":       q.get("alternate_answer"),
        "extracted_answer":       v.get("extracted_answer"),
        "is_correct":             str(is_correct_val) if is_correct_val is not None else "",
        "success":                str(success_val) if success_val is not None else "",
        "processing_time_s":      q.get("processing_time"),
        "extraction_method":      v.get("extraction_method"),
        "extraction_confidence":  v.get("extraction_confidence"),
        "match_type":             v.get("match_type"),
        "gpu_util_avg_pct":       sm.get("gpu_util_avg_percent"),
        "gpu_vram_avg_mb":        sm.get("gpu_vram_avg_mb"),
        "gpu_temp_avg_c":         sm.get("gpu_temp_avg_c"),
        "gpu_power_avg_w":        sm.get("gpu_power_avg_w"),
        "energy_wh":              sm.get("energy_estimate_wh"),
        "total_duration_s":       om.get("total_duration_s"),
        "prompt_eval_count":      om.get("prompt_eval_count"),
        "eval_count":             om.get("eval_count"),
        "generation_speed_tps":   om.get("generation_speed_tps"),
    }


def style_header(ws, row_num, num_cols):
    """Apply header styling to a row."""
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def auto_fit(ws, min_w=10, max_w=50):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), max_w))
        ws.column_dimensions[col_letter].width = max(max_len + 3, min_w)


# ──────────────────────────────────────────────────────────────────────
# Sheet writers
# ──────────────────────────────────────────────────────────────────────

def write_config_sheet(wb, fairness_config):
    """Sheet 1: Run Configuration (static settings, mentioned once)."""
    ws = wb.create_sheet("Run Configuration")

    ws.cell(row=1, column=1, value="Setting").font = HEADER_FONT
    ws.cell(row=1, column=2, value="Value").font = HEADER_FONT
    style_header(ws, 1, 2)

    rows = []

    op = fairness_config.get("ollama_parameters", {})
    for key in ("num_ctx", "num_gpu", "num_thread", "num_predict",
                "temperature", "top_p", "top_k"):
        rows.append((f"ollama_parameters.{key}", op.get(key)))

    overrides = fairness_config.get("model_overrides_config", {})
    for mname, settings in overrides.items():
        for key, val in settings.items():
            rows.append((f"model_overrides.{mname}.{key}", val))

    cd = fairness_config.get("cooldown_config", {})
    for key in ("between_models_seconds", "between_questions_seconds",
                "gpu_temp_threshold_c", "gpu_temp_wait_max_seconds"):
        rows.append((f"cooldown.{key}", cd.get(key)))

    wu = fairness_config.get("warmup_config", {})
    for key in ("enabled", "warmup_prompt"):
        rows.append((f"warmup.{key}", wu.get(key)))

    svc = fairness_config.get("system_validation_config", {})
    for key in ("enabled", "max_gpu_util_percent", "max_cpu_util_percent",
                "check_no_other_models_loaded", "validation_sample_seconds"):
        rows.append((f"system_validation.{key}", svc.get(key)))

    for i, (setting, value) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=setting).font = DATA_FONT
        cell = ws.cell(row=i, column=2)
        if isinstance(value, bool):
            cell.value = str(value)
        else:
            cell.value = value
        cell.font = DATA_FONT

    auto_fit(ws)


def write_trial_summary_sheet(wb, trial_rows):
    """Sheet 2: Trial Summary (27 rows - one per trial slot)."""
    ws = wb.create_sheet("Trial Summary")

    cols = [
        "dataset", "model", "trial_number", "status", "date_run",
        "source_file", "questions_total", "questions_answered",
        "questions_correct", "percent_correct", "successful_responses",
        "failed_responses", "total_time_s", "avg_time_per_question_s",
        "system_validation_passed", "warmup_response_time_s",
        "warmup_model_load_s",
    ]

    for ci, col_name in enumerate(cols, 1):
        ws.cell(row=1, column=ci, value=col_name)
    style_header(ws, 1, len(cols))

    for ri, trial in enumerate(trial_rows, start=2):
        is_placeholder = trial.get("status") == "PLACEHOLDER"
        for ci, col_name in enumerate(cols, 1):
            val = trial.get(col_name)
            cell = ws.cell(row=ri, column=ci)
            if isinstance(val, bool):
                cell.value = str(val)
            else:
                cell.value = val
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if is_placeholder:
                cell.fill = PLACEHOLDER_FILL

    auto_fit(ws)


def write_question_results_sheet(wb, all_rows):
    """Sheet 3: Question Results (one row per question per completed trial)."""
    ws = wb.create_sheet("Question Results")

    # Header
    for ci, (col_name, _, _) in enumerate(QR_COLUMNS, 1):
        ws.cell(row=1, column=ci, value=col_name)
    style_header(ws, 1, len(QR_COLUMNS))

    # Data rows
    for ri, row_data in enumerate(all_rows, start=2):
        for ci, (col_name, _, _) in enumerate(QR_COLUMNS, 1):
            val = row_data.get(col_name)
            cell = ws.cell(row=ri, column=ci)
            cell.value = val
            cell.font = DATA_FONT

    # Freeze the header row and first 3 columns (dataset/model/trial)
    ws.freeze_panes = "D2"

    # Auto-filter on all columns
    ws.auto_filter.ref = f"A1:{get_column_letter(len(QR_COLUMNS))}{len(all_rows) + 1}"

    auto_fit(ws)


def write_data_dictionary_sheet(wb):
    """Sheet 4: Data Dictionary (column definitions for Question Results)."""
    ws = wb.create_sheet("Data Dictionary")

    headers = ["Column Name", "Description", "JSON Source Path"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    style_header(ws, 1, len(headers))

    for ri, (col_name, description, _) in enumerate(QR_COLUMNS, start=2):
        ws.cell(row=ri, column=1, value=col_name).font = DATA_FONT
        ws.cell(row=ri, column=2, value=description).font = DATA_FONT
        ws.cell(row=ri, column=3, value=QR_SOURCE_MAP.get(col_name, "")).font = DATA_FONT

    auto_fit(ws)


# ──────────────────────────────────────────────────────────────────────
# Verification - re-read the .xlsx and compare against source data
# ──────────────────────────────────────────────────────────────────────

def verify_output(output_path, verification_records):
    """
    Re-open the generated .xlsx and cross-check row counts and correctness
    counts against the original JSON data.  Returns True if all checks pass.
    """
    wb = load_workbook(output_path, read_only=True, data_only=True)
    ws = wb["Question Results"]

    # Tally per-trial counts from the spreadsheet
    trial_row_counts = {}
    trial_correct_counts = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        ds = str(row[0])
        mdl = str(row[1])
        tn = str(row[2])
        ic = str(row[9]) if row[9] is not None else ""

        key = (ds, mdl, tn)
        trial_row_counts[key] = trial_row_counts.get(key, 0) + 1
        if ic == "True":
            trial_correct_counts[key] = trial_correct_counts.get(key, 0) + 1

    wb.close()

    all_pass = True
    total_expected = 0

    for trial_key, expected_count, expected_correct in verification_records:
        actual_count = trial_row_counts.get(trial_key, 0)
        actual_correct = trial_correct_counts.get(trial_key, 0)
        total_expected += expected_count

        ok = (actual_count == expected_count) and (actual_correct == expected_correct)
        if not ok:
            all_pass = False

        tag = "PASS" if ok else "FAIL"
        print(f"  [{tag}] {trial_key[0]} / {trial_key[1]} T{trial_key[2]}: "
              f"rows={actual_count}/{expected_count}, "
              f"correct={actual_correct}/{expected_correct}")

    total_actual = sum(trial_row_counts.values())
    total_ok = total_actual == total_expected
    if not total_ok:
        all_pass = False
    print(f"\n  Total data rows: {total_actual}  (expected {total_expected})  "
          f"{'PASS' if total_ok else 'FAIL'}")

    return all_pass


# ──────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("LLM Math Benchmark  -  Master Results Generator (.xlsx)")
    print("=" * 60)

    # ------------------------------------------------------------------
    # 1. Load JSON files and build per-trial data
    # ------------------------------------------------------------------
    print("\nLoading result files ...")
    trial_summary_rows = []
    all_question_rows = []
    verification_records = []
    fairness_config = None

    for dataset, model, trial_num, filepath in TRIAL_REGISTRY:

        if filepath is None:
            trial_summary_rows.append({
                "dataset": dataset,
                "model": model,
                "trial_number": trial_num,
                "status": "PLACEHOLDER",
                "date_run": None,
                "source_file": None,
                "questions_total": None,
                "questions_answered": None,
                "questions_correct": None,
                "percent_correct": None,
                "successful_responses": None,
                "failed_responses": None,
                "total_time_s": None,
                "avg_time_per_question_s": None,
                "system_validation_passed": None,
                "warmup_response_time_s": None,
                "warmup_model_load_s": None,
            })
            continue

        data = load_json(filepath)
        summary = data.get("summary", {})
        metadata = data.get("metadata", {})
        fc = metadata.get("fairness_config", {})
        results = data.get("results", [])

        if fairness_config is None:
            fairness_config = fc

        warmup = fc.get("warmup_result") or {}
        sysval = fc.get("system_validation_result") or {}

        questions_answered = summary.get("questions_answered", len(results))
        questions_correct = summary.get("questions_correct", 0)

        if "advanced_probability" in filepath:
            q_total = 1000
        elif "calculus1" in filepath:
            q_total = 900
        else:
            q_total = 644

        trial_summary_rows.append({
            "dataset": dataset,
            "model": model,
            "trial_number": trial_num,
            "status": "COMPLETE",
            "date_run": metadata.get("created", ""),
            "source_file": filepath,
            "questions_total": q_total,
            "questions_answered": questions_answered,
            "questions_correct": questions_correct,
            "percent_correct": summary.get("percent_correct"),
            "successful_responses": metadata.get("successful_responses"),
            "failed_responses": metadata.get("failed_responses"),
            "total_time_s": summary.get("total_time_seconds"),
            "avg_time_per_question_s": summary.get("average_time_per_question_seconds"),
            "system_validation_passed":
                str(sysval.get("passed")) if sysval.get("passed") is not None else "",
            "warmup_response_time_s": warmup.get("warmup_response_time_s"),
            "warmup_model_load_s": warmup.get("model_load_duration_s"),
        })

        row_count = 0
        correct_count = 0
        for q in results:
            row = extract_question_row(dataset, model, trial_num, q)
            all_question_rows.append(row)
            row_count += 1
            if q.get("is_correct"):
                correct_count += 1

        print(f"  {dataset} / {model} Trial {trial_num}: "
              f"{row_count} questions, {correct_count} correct")

        if row_count != questions_answered:
            print(f"    WARNING: JSON summary says {questions_answered} answered "
                  f"but found {row_count} result objects!")
        if correct_count != questions_correct:
            print(f"    WARNING: JSON summary says {questions_correct} correct "
                  f"but counted {correct_count} is_correct=true!")

        verification_records.append(
            ((dataset, model, str(trial_num)), row_count, correct_count)
        )

    print(f"\nTotal question rows to write: {len(all_question_rows)}")
    print(f"Trial summary rows: {len(trial_summary_rows)} "
          f"({sum(1 for t in trial_summary_rows if t['status'] == 'COMPLETE')} complete, "
          f"{sum(1 for t in trial_summary_rows if t['status'] == 'PLACEHOLDER')} placeholder)")

    # ------------------------------------------------------------------
    # 2. Build Excel workbook
    # ------------------------------------------------------------------
    print("\nGenerating .xlsx workbook ...")
    wb = Workbook()

    write_config_sheet(wb, fairness_config)
    write_trial_summary_sheet(wb, trial_summary_rows)
    write_question_results_sheet(wb, all_question_rows)
    write_data_dictionary_sheet(wb)

    # Remove the default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ------------------------------------------------------------------
    # 3. Save to disk
    # ------------------------------------------------------------------
    output_path = os.path.join(PROJECT_ROOT, "master_results.xlsx")
    wb.save(output_path)
    wb.close()

    size_mb = os.path.getsize(output_path) / (1024 * 1024)
    print(f"Written to: {output_path}")
    print(f"File size:  {size_mb:.1f} MB")

    # ------------------------------------------------------------------
    # 4. Post-write verification (re-read the .xlsx)
    # ------------------------------------------------------------------
    print("\n--- Verification (re-reading generated .xlsx) ---")
    passed = verify_output(output_path, verification_records)

    if passed:
        print("\nAll verification checks PASSED.")
    else:
        print("\nERROR: Some verification checks FAILED!")
        sys.exit(1)

    print("\nDone!")


if __name__ == "__main__":
    main()
