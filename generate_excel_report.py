#!/usr/bin/env python3
"""
Generate professional Excel spreadsheet analyzing LLM math benchmark results.
Compares Gemma3:4b, Phi3:3.8b, and Qwen3:4b across Calculus I,
Advanced Probability & Statistics, and Grade 8 Math categories with subcategory breakdowns.
"""

import json
import os
import sys
import xml.etree.ElementTree as ET
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))

RESULT_FILES = {
    ("Gemma3:4b", "Calculus I"): "results/Gemma/calculus1_problems_gemma3_4b_02-20-26T03-14-07.json",
    ("Gemma3:4b", "Advanced Probability & Statistics"): "results/Gemma/advanced_probability_statistics_problems_gemma3_4b_02-20-26T09-37-45.json",
    ("Gemma3:4b", "Grade 8 Math"): "results/Gemma/grade_8_math_problems_gemma3_4b_02-20-26T13-46-16.json",
    ("Phi3:3.8b", "Calculus I"): "results/Phi/calculus1_problems_phi3_3_8b_02-19-26T23-08-26.json",
    ("Phi3:3.8b", "Advanced Probability & Statistics"): "results/Phi/advanced_probability_statistics_problems_phi3_3_8b_02-20-26T00-27-05.json",
    ("Phi3:3.8b", "Grade 8 Math"): "results/Phi/grade_8_math_problems_phi3_3_8b_02-20-26T02-01-22.json",
    ("Qwen3:4b", "Calculus I"): "results/Qwen/calculus1_problems_qwen3_4b_02-21-26T11-13-03.json",
    ("Qwen3:4b", "Advanced Probability & Statistics"): "results/Qwen/advanced_probability_statistics_problems_qwen3_4b_02-20-26T17-44-52.json",
    ("Qwen3:4b", "Grade 8 Math"): "results/Qwen/grade_8_math_problems_qwen3_4b_02-21-26T01-00-40.json",
}

XML_FILES = {
    "Calculus I": "calculus1_problems.xml",
    "Advanced Probability & Statistics": "advanced_probability_statistics_problems.xml",
    "Grade 8 Math": "grade_8_math_problems.xml",
}

MODELS = ["Gemma3:4b", "Phi3:3.8b", "Qwen3:4b"]
CATEGORIES = ["Advanced Probability & Statistics", "Calculus I", "Grade 8 Math"]

# Style constants
HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
SUBHEADER_FONT = Font(name='Calibri', bold=True, size=10, color='2F5496')
SUBHEADER_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
DATA_FONT = Font(name='Calibri', size=10)
TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='2F5496')
SUBTITLE_FONT = Font(name='Calibri', italic=True, size=10, color='808080')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
GREEN_FONT = Font(name='Calibri', size=10, color='006100')
YELLOW_FONT = Font(name='Calibri', size=10, color='9C6500')
RED_FONT = Font(name='Calibri', size=10, color='9C0006')

PCT_FMT = '0.0%'
NUM_FMT = '#,##0'
DEC2_FMT = '0.00'
TIME_FMT = '#,##0.00'


# ---------------------------------------------------------------------------
# Helpers – styling
# ---------------------------------------------------------------------------
def style_header_row(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if fmt:
        cell.number_format = fmt
    return cell


def write_cell(ws, row, col, value, fmt=None, font=None, fill=None, alignment=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = THIN_BORDER
    cell.font = font or DATA_FONT
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    cell.alignment = alignment or Alignment(horizontal='center', vertical='center')
    return cell


def apply_accuracy_cond_fmt(ws, col_letter, start_row, end_row):
    rng = f'{col_letter}{start_row}:{col_letter}{end_row}'
    ws.conditional_formatting.add(rng, CellIsRule(
        operator='greaterThan', formula=['0.70'], fill=GREEN_FILL, font=GREEN_FONT))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator='between', formula=['0.40', '0.70'], fill=YELLOW_FILL, font=YELLOW_FONT))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator='lessThan', formula=['0.40'], fill=RED_FILL, font=RED_FONT))


def auto_fit_columns(ws, min_width=10, max_width=45):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min(max_len + 3, max_width), min_width)


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------
def load_result_file(filepath):
    abs_path = os.path.join(PROJECT_ROOT, filepath)
    try:
        with open(abs_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except json.JSONDecodeError as e:
        print(f"  WARNING: JSON decode error in {filepath}: {e}")
        try:
            with open(abs_path, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read()
            # Try to recover by finding last complete result object
            last_bracket = content.rfind(']')
            if last_bracket > 0:
                truncated = content[:last_bracket + 1] + '}'
                return json.loads(truncated)
        except Exception as e2:
            print(f"  Recovery failed: {e2}")
        return None


def load_xml_directions(xml_file):
    abs_path = os.path.join(PROJECT_ROOT, xml_file)
    tree = ET.parse(abs_path)
    root = tree.getroot()
    mapping = {}
    for row in root.findall('row'):
        pn = row.find('ProblemNumber')
        dr = row.find('Directions')
        if pn is not None and pn.text and dr is not None and dr.text:
            mapping[pn.text.strip()] = dr.text.strip()
    return mapping


def build_dataframe(result_data, model, category, directions_map):
    rows = []
    for r in result_data.get('results', []):
        v = r.get('verification') or {}
        rows.append({
            'model': model,
            'category_label': category,
            'question_id': str(r.get('question_id', '')),
            'subcategory': r.get('category', 'unknown'),
            'directions': directions_map.get(str(r.get('question_id', '')), 'Unknown'),
            'is_correct': bool(r.get('is_correct', False)),
            'processing_time': float(r.get('processing_time', 0) or 0),
            'success': bool(r.get('success', False)),
            'match_type': v.get('match_type', 'unknown'),
            'extraction_confidence': float(v.get('extraction_confidence', 0) or 0),
            'comparison_confidence': float(v.get('comparison_confidence', 0) or 0),
            'verification_status': v.get('verification_status', 'unknown'),
        })
    return pd.DataFrame(rows)


def compute_metrics(df):
    n = len(df)
    if n == 0:
        return {'count': 0, 'correct': 0, 'accuracy': 0.0,
                'avg_time': 0.0, 'median_time': 0.0, 'std_time': 0.0,
                'min_time': 0.0, 'max_time': 0.0,
                'p25_time': 0.0, 'p75_time': 0.0, 'p95_time': 0.0,
                'total_time': 0.0}
    correct = int(df['is_correct'].sum())
    t = df['processing_time']
    return {
        'count': n,
        'correct': correct,
        'accuracy': correct / n if n else 0.0,
        'avg_time': float(t.mean()),
        'median_time': float(t.median()),
        'std_time': float(t.std()) if n > 1 else 0.0,
        'min_time': float(t.min()),
        'max_time': float(t.max()),
        'p25_time': float(t.quantile(0.25)),
        'p75_time': float(t.quantile(0.75)),
        'p95_time': float(t.quantile(0.95)),
        'total_time': float(t.sum()),
    }


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------
def write_executive_summary(wb, df, warnings):
    ws = wb.create_sheet("Executive Summary", 0)
    num_models = len(MODELS)
    last_col = 1 + num_models

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = ws.cell(row=1, column=1, value="LLM Math Benchmark — Executive Summary")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='left')

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = SUBTITLE_FONT

    # Overall model comparison table
    row = 4
    headers = ["Metric"] + MODELS
    for ci, h in enumerate(headers, 1):
        write_cell(ws, row, ci, h, font=HEADER_FONT, fill=HEADER_FILL)
    style_header_row(ws, row, 1, len(headers))

    metrics_rows = []
    for model in MODELS:
        m = compute_metrics(df[df['model'] == model])
        metrics_rows.append(m)

    labels = [
        ("Total Questions Attempted", 'count', NUM_FMT),
        ("Total Correct", 'correct', NUM_FMT),
        ("Overall Accuracy", 'accuracy', PCT_FMT),
        ("Total Processing Time (s)", 'total_time', TIME_FMT),
        ("Avg Time Per Question (s)", 'avg_time', DEC2_FMT),
        ("Median Time Per Question (s)", 'median_time', DEC2_FMT),
    ]

    for i, (label, key, fmt) in enumerate(labels):
        r = row + 1 + i
        write_cell(ws, r, 1, label, font=SUBHEADER_FONT,
                   alignment=Alignment(horizontal='left', vertical='center'))
        for mi, m in enumerate(metrics_rows):
            write_cell(ws, r, 2 + mi, m[key], fmt=fmt)

    # Accuracy conditional formatting for each model column
    acc_row = row + 3  # the Accuracy row
    for mi in range(num_models):
        col_letter = get_column_letter(2 + mi)
        apply_accuracy_cond_fmt(ws, col_letter, acc_row, acc_row)

    # Per-category quick view
    row = row + len(labels) + 3
    cat_last_col = 1 + num_models * 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cat_last_col)
    ws.cell(row=row, column=1, value="Performance by Category").font = Font(
        name='Calibri', bold=True, size=12, color='2F5496')

    row += 1
    cat_headers = ["Category"]
    for model in MODELS:
        cat_headers += [f"{model} Questions", f"{model} Accuracy"]
    for ci, h in enumerate(cat_headers, 1):
        write_cell(ws, row, ci, h, font=HEADER_FONT, fill=HEADER_FILL)
    style_header_row(ws, row, 1, len(cat_headers))

    for i, cat in enumerate(CATEGORIES):
        r = row + 1 + i
        write_cell(ws, r, 1, cat, font=SUBHEADER_FONT,
                   alignment=Alignment(horizontal='left', vertical='center'))
        for mi, model in enumerate(MODELS):
            m = compute_metrics(df[(df['model'] == model) & (df['category_label'] == cat)])
            write_cell(ws, r, 2 + mi * 2, m['count'], fmt=NUM_FMT)
            write_cell(ws, r, 3 + mi * 2, m['accuracy'], fmt=PCT_FMT)

    cat_data_start = row + 1
    cat_data_end = row + len(CATEGORIES)
    for mi in range(num_models):
        acc_col = get_column_letter(3 + mi * 2)
        apply_accuracy_cond_fmt(ws, acc_col, cat_data_start, cat_data_end)

    # Warnings / notes
    if warnings:
        row = cat_data_end + 2
        ws.cell(row=row, column=1, value="Data Notes:").font = Font(
            name='Calibri', bold=True, size=10, color='C00000')
        for wi, w in enumerate(warnings):
            ws.cell(row=row + 1 + wi, column=1, value=f"  • {w}").font = Font(
                name='Calibri', size=9, color='808080')

    auto_fit_columns(ws)


def write_category_comparison(wb, df):
    ws = wb.create_sheet("Category Comparison")

    ws.merge_cells('A1:I1')
    ws.cell(row=1, column=1, value="Category-Level Comparison").font = TITLE_FONT

    headers = ["Category", "Model", "Questions", "Correct", "Incorrect",
               "Accuracy %", "Avg Time (s)", "Median Time (s)",
               "Min Time (s)", "Max Time (s)"]
    row = 3
    for ci, h in enumerate(headers, 1):
        write_cell(ws, row, ci, h, font=HEADER_FONT, fill=HEADER_FILL)
    style_header_row(ws, row, 1, len(headers))

    data_start = row + 1
    r = data_start
    for cat in CATEGORIES:
        for model in MODELS:
            m = compute_metrics(df[(df['model'] == model) & (df['category_label'] == cat)])
            write_cell(ws, r, 1, cat, font=SUBHEADER_FONT,
                       alignment=Alignment(horizontal='left', vertical='center'))
            write_cell(ws, r, 2, model)
            write_cell(ws, r, 3, m['count'], fmt=NUM_FMT)
            write_cell(ws, r, 4, m['correct'], fmt=NUM_FMT)
            write_cell(ws, r, 5, m['count'] - m['correct'], fmt=NUM_FMT)
            write_cell(ws, r, 6, m['accuracy'], fmt=PCT_FMT)
            write_cell(ws, r, 7, m['avg_time'], fmt=DEC2_FMT)
            write_cell(ws, r, 8, m['median_time'], fmt=DEC2_FMT)
            write_cell(ws, r, 9, m['min_time'], fmt=DEC2_FMT)
            write_cell(ws, r, 10, m['max_time'], fmt=DEC2_FMT)
            r += 1

    apply_accuracy_cond_fmt(ws, 'F', data_start, r - 1)
    auto_fit_columns(ws)


def _write_subcategory_sheet(wb, df, sheet_name, title, category_label, group_col):
    """Generic helper for subcategory breakdown sheets."""
    ws = wb.create_sheet(sheet_name)
    num_models = len(MODELS)
    total_cols = 1 + num_models * 4

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT

    # Build headers dynamically
    headers = ["Topic / Subcategory"]
    for model in MODELS:
        headers += [f"{model} Count", f"{model} Correct", f"{model} Accuracy", f"{model} Avg Time (s)"]

    row = 3
    for ci, h in enumerate(headers, 1):
        write_cell(ws, row, ci, h, font=HEADER_FONT, fill=HEADER_FILL)
    style_header_row(ws, row, 1, len(headers))

    cat_df = df[df['category_label'] == category_label].copy()
    subcats = sorted(cat_df[group_col].unique(), key=lambda x: x if x != 'unknown' else 'zzz')

    data_start = row + 1
    r = data_start
    totals = {model: {'count': 0, 'correct': 0, 'total_time': 0.0} for model in MODELS}

    for sub in subcats:
        write_cell(ws, r, 1, sub, font=DATA_FONT,
                   alignment=Alignment(horizontal='left', vertical='center', wrap_text=True))
        for mi, model in enumerate(MODELS):
            sub_df = cat_df[(cat_df['model'] == model) & (cat_df[group_col] == sub)]
            m = compute_metrics(sub_df)
            col_offset = 2 + mi * 4
            write_cell(ws, r, col_offset, m['count'], fmt=NUM_FMT)
            write_cell(ws, r, col_offset + 1, m['correct'], fmt=NUM_FMT)
            write_cell(ws, r, col_offset + 2, m['accuracy'], fmt=PCT_FMT)
            write_cell(ws, r, col_offset + 3, m['avg_time'], fmt=DEC2_FMT)
            totals[model]['count'] += m['count']
            totals[model]['correct'] += m['correct']
            totals[model]['total_time'] += m['total_time']
        r += 1

    # Totals row
    write_cell(ws, r, 1, "TOTAL", font=Font(name='Calibri', bold=True, size=10),
               fill=SUBHEADER_FILL, alignment=Alignment(horizontal='left', vertical='center'))
    for mi, model in enumerate(MODELS):
        col_offset = 2 + mi * 4
        tc = totals[model]['count']
        tcorr = totals[model]['correct']
        write_cell(ws, r, col_offset, tc, fmt=NUM_FMT, fill=SUBHEADER_FILL,
                   font=Font(name='Calibri', bold=True, size=10))
        write_cell(ws, r, col_offset + 1, tcorr, fmt=NUM_FMT, fill=SUBHEADER_FILL,
                   font=Font(name='Calibri', bold=True, size=10))
        write_cell(ws, r, col_offset + 2, tcorr / tc if tc else 0, fmt=PCT_FMT,
                   fill=SUBHEADER_FILL, font=Font(name='Calibri', bold=True, size=10))
        avg_t = totals[model]['total_time'] / tc if tc else 0
        write_cell(ws, r, col_offset + 3, avg_t, fmt=DEC2_FMT, fill=SUBHEADER_FILL,
                   font=Font(name='Calibri', bold=True, size=10))

    # Conditional formatting on accuracy columns for each model
    for mi in range(num_models):
        acc_col = get_column_letter(4 + mi * 4)  # col 4, 8, 12, ...
        apply_accuracy_cond_fmt(ws, acc_col, data_start, r)

    auto_fit_columns(ws)


def write_grade8_breakdown(wb, df):
    _write_subcategory_sheet(
        wb, df,
        sheet_name="Grade 8 Math Breakdown",
        title="Grade 8 Math — Subcategory Breakdown",
        category_label="Grade 8 Math",
        group_col="subcategory",
    )


def write_calculus_breakdown(wb, df):
    _write_subcategory_sheet(
        wb, df,
        sheet_name="Calculus I Breakdown",
        title="Calculus I — Topic Breakdown (by Directions)",
        category_label="Calculus I",
        group_col="directions",
    )


def write_stats_breakdown(wb, df):
    _write_subcategory_sheet(
        wb, df,
        sheet_name="Prob & Stats Breakdown",
        title="Advanced Probability & Statistics — Topic Breakdown (by Directions)",
        category_label="Advanced Probability & Statistics",
        group_col="directions",
    )


def write_verification_analysis(wb, df):
    ws = wb.create_sheet("Verification Analysis")
    num_models = len(MODELS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + num_models * 2)
    ws.cell(row=1, column=1, value="Verification & Match Analysis").font = TITLE_FONT

    # --- Sub-table 1: Overall Match Type Distribution ---
    row = 3
    ws.cell(row=row, column=1, value="Match Type Distribution (Overall)").font = Font(
        name='Calibri', bold=True, size=11, color='2F5496')
    row += 1
    mt_headers = ["Match Type"]
    for model in MODELS:
        mt_headers += [f"{model} Count", f"{model} %"]
    for ci, h in enumerate(mt_headers, 1):
        write_cell(ws, row, ci, h, font=HEADER_FONT, fill=HEADER_FILL)
    style_header_row(ws, row, 1, len(mt_headers))

    match_types = sorted(df['match_type'].unique())
    r = row + 1
    for mt in match_types:
        write_cell(ws, r, 1, mt, alignment=Alignment(horizontal='left', vertical='center'))
        for mi, model in enumerate(MODELS):
            model_df = df[df['model'] == model]
            cnt = int((model_df['match_type'] == mt).sum())
            pct = cnt / len(model_df) if len(model_df) else 0
            write_cell(ws, r, 2 + mi * 2, cnt, fmt=NUM_FMT)
            write_cell(ws, r, 3 + mi * 2, pct, fmt=PCT_FMT)
        r += 1

    # --- Sub-table 2: Match Type by Category ---
    r += 1
    ws.cell(row=r, column=1, value="Match Type Distribution by Category").font = Font(
        name='Calibri', bold=True, size=11, color='2F5496')
    r += 1
    mt_cat_headers = ["Category", "Model", "exact", "equivalent", "no_match", "Other", "Total"]
    for ci, h in enumerate(mt_cat_headers, 1):
        write_cell(ws, r, ci, h, font=HEADER_FONT, fill=HEADER_FILL)
    style_header_row(ws, r, 1, len(mt_cat_headers))

    r += 1
    for cat in CATEGORIES:
        for model in MODELS:
            subset = df[(df['model'] == model) & (df['category_label'] == cat)]
            mt_counts = subset['match_type'].value_counts()
            exact_n = int(mt_counts.get('exact', 0))
            equiv_n = int(mt_counts.get('equivalent', 0))
            nomatch_n = int(mt_counts.get('no_match', 0))
            other_n = int(len(subset) - exact_n - equiv_n - nomatch_n)
            write_cell(ws, r, 1, cat, alignment=Alignment(horizontal='left', vertical='center'))
            write_cell(ws, r, 2, model)
            write_cell(ws, r, 3, exact_n, fmt=NUM_FMT)
            write_cell(ws, r, 4, equiv_n, fmt=NUM_FMT)
            write_cell(ws, r, 5, nomatch_n, fmt=NUM_FMT)
            write_cell(ws, r, 6, other_n, fmt=NUM_FMT)
            write_cell(ws, r, 7, len(subset), fmt=NUM_FMT)
            r += 1

    # --- Sub-table 3: Extraction Confidence Stats ---
    r += 1
    ws.cell(row=r, column=1, value="Extraction & Comparison Confidence").font = Font(
        name='Calibri', bold=True, size=11, color='2F5496')
    r += 1
    conf_headers = ["Metric"] + MODELS
    for ci, h in enumerate(conf_headers, 1):
        write_cell(ws, r, ci, h, font=HEADER_FONT, fill=HEADER_FILL)
    style_header_row(ws, r, 1, len(conf_headers))

    r += 1
    for label, col_name in [("Mean Extraction Confidence", "extraction_confidence"),
                             ("Mean Comparison Confidence", "comparison_confidence")]:
        write_cell(ws, r, 1, label, alignment=Alignment(horizontal='left', vertical='center'))
        for mi, model in enumerate(MODELS):
            val = df[df['model'] == model][col_name].mean()
            write_cell(ws, r, 2 + mi, val, fmt=DEC2_FMT)
        r += 1

    # --- Sub-table 4: Verification Status ---
    r += 1
    ws.cell(row=r, column=1, value="Verification Status Distribution").font = Font(
        name='Calibri', bold=True, size=11, color='2F5496')
    r += 1
    vs_headers = ["Status"]
    for model in MODELS:
        vs_headers += [f"{model} Count", f"{model} %"]
    for ci, h in enumerate(vs_headers, 1):
        write_cell(ws, r, ci, h, font=HEADER_FONT, fill=HEADER_FILL)
    style_header_row(ws, r, 1, len(vs_headers))

    r += 1
    statuses = sorted(df['verification_status'].unique())
    for vs in statuses:
        write_cell(ws, r, 1, vs, alignment=Alignment(horizontal='left', vertical='center'))
        for mi, model in enumerate(MODELS):
            model_df = df[df['model'] == model]
            cnt = int((model_df['verification_status'] == vs).sum())
            pct = cnt / len(model_df) if len(model_df) else 0
            write_cell(ws, r, 2 + mi * 2, cnt, fmt=NUM_FMT)
            write_cell(ws, r, 3 + mi * 2, pct, fmt=PCT_FMT)
        r += 1

    auto_fit_columns(ws)


def write_timing_analysis(wb, df):
    ws = wb.create_sheet("Timing Analysis")

    ws.merge_cells('A1:K1')
    ws.cell(row=1, column=1, value="Processing Time Analysis").font = TITLE_FONT

    # --- Table 1: Per-category per-model timing ---
    row = 3
    headers = ["Category", "Model", "Count", "Mean (s)", "Median (s)", "Std Dev (s)",
               "P25 (s)", "P75 (s)", "P95 (s)", "Min (s)", "Max (s)"]
    for ci, h in enumerate(headers, 1):
        write_cell(ws, row, ci, h, font=HEADER_FONT, fill=HEADER_FILL)
    style_header_row(ws, row, 1, len(headers))

    r = row + 1
    for cat in CATEGORIES:
        for model in MODELS:
            m = compute_metrics(df[(df['model'] == model) & (df['category_label'] == cat)])
            write_cell(ws, r, 1, cat, alignment=Alignment(horizontal='left', vertical='center'))
            write_cell(ws, r, 2, model)
            write_cell(ws, r, 3, m['count'], fmt=NUM_FMT)
            write_cell(ws, r, 4, m['avg_time'], fmt=DEC2_FMT)
            write_cell(ws, r, 5, m['median_time'], fmt=DEC2_FMT)
            write_cell(ws, r, 6, m['std_time'], fmt=DEC2_FMT)
            write_cell(ws, r, 7, m['p25_time'], fmt=DEC2_FMT)
            write_cell(ws, r, 8, m['p75_time'], fmt=DEC2_FMT)
            write_cell(ws, r, 9, m['p95_time'], fmt=DEC2_FMT)
            write_cell(ws, r, 10, m['min_time'], fmt=DEC2_FMT)
            write_cell(ws, r, 11, m['max_time'], fmt=DEC2_FMT)
            r += 1

    # --- Table 2: Overall Model Timing ---
    r += 1
    ws.cell(row=r, column=1, value="Overall Model Timing Summary").font = Font(
        name='Calibri', bold=True, size=11, color='2F5496')
    r += 1
    ov_headers = ["Model", "Total Questions", "Total Time (s)", "Total Time (min)",
                  "Avg Time (s)", "Median Time (s)"]
    for ci, h in enumerate(ov_headers, 1):
        write_cell(ws, r, ci, h, font=HEADER_FONT, fill=HEADER_FILL)
    style_header_row(ws, r, 1, len(ov_headers))

    r += 1
    for model in MODELS:
        m = compute_metrics(df[df['model'] == model])
        write_cell(ws, r, 1, model, alignment=Alignment(horizontal='left', vertical='center'))
        write_cell(ws, r, 2, m['count'], fmt=NUM_FMT)
        write_cell(ws, r, 3, m['total_time'], fmt=TIME_FMT)
        write_cell(ws, r, 4, m['total_time'] / 60, fmt=DEC2_FMT)
        write_cell(ws, r, 5, m['avg_time'], fmt=DEC2_FMT)
        write_cell(ws, r, 6, m['median_time'], fmt=DEC2_FMT)
        r += 1

    auto_fit_columns(ws)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("=" * 60)
    print("LLM Math Benchmark — Excel Report Generator")
    print("=" * 60)

    # Step 1: Load XML directions
    print("\nLoading XML direction mappings...")
    xml_directions = {}
    for cat, xml_file in XML_FILES.items():
        xml_directions[cat] = load_xml_directions(xml_file)
        print(f"  {cat}: {len(xml_directions[cat])} directions loaded")

    # Step 2: Load result files
    print("\nLoading result files...")
    all_dfs = []
    warnings = []
    for (model, category), filepath in RESULT_FILES.items():
        data = load_result_file(filepath)
        if data is None:
            msg = f"FAILED to load {filepath}"
            print(f"  ERROR: {msg}")
            warnings.append(msg)
            continue

        results = data.get('results', [])
        summary = data.get('summary', {})
        actual_count = len(results)
        actual_correct = sum(1 for r in results if r.get('is_correct'))
        summary_count = summary.get('questions_answered', 0)

        if summary_count != actual_count:
            msg = (f"{model} / {category}: Summary says {summary_count} answered, "
                   f"but {actual_count} results exist. Using actual count ({actual_correct} correct).")
            print(f"  NOTE: {msg}")
            warnings.append(msg)
        else:
            print(f"  {model} / {category}: {actual_count} results, {actual_correct} correct "
                  f"({actual_correct/actual_count*100:.1f}%)")

        directions_map = xml_directions.get(category, {})
        frame = build_dataframe(data, model, category, directions_map)
        all_dfs.append(frame)

    if not all_dfs:
        print("\nERROR: No data loaded. Exiting.")
        sys.exit(1)

    combined = pd.concat(all_dfs, ignore_index=True)
    print(f"\nTotal records: {len(combined)}")

    # Step 3: Validation checks
    print("\n--- Validation ---")
    for model in MODELS:
        for cat in CATEGORIES:
            subset = combined[(combined['model'] == model) & (combined['category_label'] == cat)]
            unk_dirs = (subset['directions'] == 'Unknown').sum()
            print(f"  {model} / {cat}: {len(subset)} records, "
                  f"{int(subset['is_correct'].sum())} correct, "
                  f"{unk_dirs} unmatched directions")

    # Step 4: Generate Excel
    print("\nGenerating Excel workbook...")
    wb = Workbook()

    write_executive_summary(wb, combined, warnings)
    write_category_comparison(wb, combined)
    write_grade8_breakdown(wb, combined)
    write_calculus_breakdown(wb, combined)
    write_stats_breakdown(wb, combined)
    write_verification_analysis(wb, combined)
    write_timing_analysis(wb, combined)

    # Remove default empty sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    output_path = os.path.join(PROJECT_ROOT, 'analysis_results.xlsx')
    wb.save(output_path)
    print(f"\nSpreadsheet saved to: {output_path}")

    # Step 5: Final verification
    print("\n--- Final Verification ---")
    for model in MODELS:
        m = compute_metrics(combined[combined['model'] == model])
        print(f"  {model}: {m['count']} total, {m['correct']} correct, "
              f"{m['accuracy']*100:.1f}% accuracy, {m['total_time']:.0f}s total time")

    print(f"\nSheets: {wb.sheetnames}")
    print("Done!")


if __name__ == '__main__':
    main()
